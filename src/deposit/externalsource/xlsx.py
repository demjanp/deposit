from deposit.externalsource.abstract_externalsource import AbstractExternalsource

from openpyxl.styles import (Font)
from openpyxl import (load_workbook, Workbook)

class XLSX(AbstractExternalsource):
	
	def __init__(self):
		
		AbstractExternalsource.__init__(self)
		
		self._columns = {} # {sheet: {column_idx: name, ...}, ...}
		self._data = {} # {sheet: [row, ...], ...}; row = {column_idx: value, ...}
	
	def sheets(self):
		
		return list(self._columns.keys())
	
	def column_count(self, sheet):
		
		if sheet in self._columns:
			return len(self._columns[sheet])
		return 0
	
	def row_count(self, sheet):
		
		if sheet not in self._data:
			return 0
		return len(self._data[sheet])
	
	def column_name(self, sheet, column_idx):
		
		if sheet not in self._data:
			return None
		if column_idx not in self._columns[sheet]:
			return None
		return self._columns[sheet][column_idx]
	
	def data(self, sheet, row_idx, column_idx):
		
		if sheet not in self._data:
			return None
		if row_idx > len(self._data[sheet]) - 1:
			return None
		if column_idx not in self._data[sheet][row_idx]:
			return None
		return self._data[sheet][row_idx][column_idx]
	
	def save(self, get_header, get_item, n_rows, n_cols, path = None, url = None, *args, **kwargs):
		# get_header(col, user_role = False) = column_name or (class_name, descriptor_name) if user_role = True
		# get_item(row, col) = deposit_gui.QueryItem
		
		self.set_path(path, url)
		path = self.get_path()
		if path is None:
			return False
		
		wb = Workbook()
		wb.guess_types = True
		ws = wb.active
		ws.append(self.get_header_data(get_header, n_cols))
		for col in range(n_cols):
			ws.cell(row = 1, column = col + 1).font = Font(bold = True)
		for row in range(n_rows):
			ws.append(self.get_row_data(row, get_item, n_cols))
		wb.save(path)
		
		return True
	
	def load(self, path = None, url = None, *args, **kwargs):
		
		self.set_path(path, url)
		path = self.get_path()
		if path is None:
			return False
		
		self._columns.clear()
		self._data.clear()
		
		wb = load_workbook(filename = path, read_only = True)
		for sheet in wb.sheetnames:
			self._columns[sheet] = {}
			self._data[sheet] = []
			ws = wb[sheet]
			idx = 0
			idx_lookup = {} # {i: idx, ...}
			for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
				value = cell.value
				if value is not None:
					value = str(value).strip()
				if value:
					self._columns[sheet][idx] = value
					idx_lookup[i] = idx
					idx += 1
			if not self._columns[sheet]:
				continue
			for row in ws.iter_rows(min_row = 2):
				self._data[sheet].append({})
				for i, cell in enumerate(row):
					value = cell.value
					if value is not None:
						value = str(value).strip()
					if value and (i in idx_lookup):
						self._data[sheet][-1][idx_lookup[i]] = value
		
		return True
		
	
		

