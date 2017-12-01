'''
	Deposit file store - XLSX
	--------------------------
	
	Created on 10. 4. 2017
	
	@author: Peter Demjan <peter.demjan@gmail.com>
	
	accessible as file.xlsx
'''

import os
from openpyxl.styles import (Font)
from openpyxl import (load_workbook, Workbook)

class XLSX(object):
	
	def __init__(self, store):
		
		self.__store = store
	
	def __getattr__(self, key):
		
		return getattr(self.__store, key)
	
	def is_xlsx(self, path):
		# return True if path is an xlsx file
		
		return (os.path.splitext(path)[-1].lower() == ".xlsx")
	
	def get_sheets(self, path):
		# return names of sheets in XLSX file
		# [name, ...]
		# TODO support remotely stored files
		
		wb = load_workbook(filename = path, read_only = True)
		return wb.get_sheet_names()
	
	def get_sheet(self, path, sheet):
		# return columns, data
		# columns: [name, ...]
		# data: [{column: value, ...}, ...]
		# TODO support remotely stored files
		
		wb = load_workbook(filename = path, read_only = True)
		ws = wb[sheet]
		data = [] # [{column: value, ...}, ...]
		columns = {} # {col_number: name, ...}
		names = {}
		for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
			if not cell.value is None:
				columns[i] = str(cell.value)
		if columns:
			for row in ws.iter_rows(min_row = 2):
				data.append({})
				for i, cell in enumerate(row):
					if i in columns:
						data[-1][i] = cell.value
		if columns:
			columns = [columns[i] for i in sorted(columns.keys())]
		else:
			columns = []
		
		return [columns, data]
	
	def write(self, path, fields = [], data = []):
		# create an XLSX file
		# fields: [name, ...]
		# data: [{field: value, ...}, ...]
		
		wb = Workbook()
		wb.guess_types = True
		ws = wb.active
		if not fields:
			for row in data:
				for field in row:
					if not field in fields:
						fields.append(field)
			fields = sorted(fields)
		ws.append(fields)
		for i in range(len(fields)):
			ws.cell(row = 1, column = i + 1).font = Font(bold = True)
		for row in data:
			ws.append([(row[field] if field in row else "") for field in fields])
		wb.save(path)
	
