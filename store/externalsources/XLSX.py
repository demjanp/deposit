from deposit.store.externalsources._ExternalSource import (ExternalSource)
from deposit.store.DLabel.DString import (DString)
from deposit.store.Conversions import (as_path)

from openpyxl.styles import (Font)
from openpyxl import (load_workbook, Workbook)

class XLSX(ExternalSource):
	
	def __init__(self, store, url):
		
		self._columns = {} # {sheet: {column_idx: name, ...}, ...}
		self._data = {} # {sheet: [row, ...], ...}; row = {column_idx: value, ...}
		
		ExternalSource.__init__(self, store, url)
	
	def load(self):
		
		self._columns.clear()
		self._data.clear()
		
		path = as_path(self.url)
		if path is None:
			return False
		
		wb = load_workbook(filename = path, read_only = True)
		for sheet in wb.sheetnames:
			self._columns[sheet] = {}
			self._data[sheet] = []
			ws = wb[sheet]
			idx = 0
			idx_lookup = {} # {i: idx, ...}
			for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
				value = cell.value
				if value is not None:
					value = str(value).strip()
				if value:
					self._columns[sheet][idx] = value
					idx_lookup[i] = idx
					idx += 1
			if not self._columns[sheet]:
				continue
			for row in ws.iter_rows(min_row = 2):
				self._data[sheet].append({})
				for i, cell in enumerate(row):
					value = cell.value
					if value is not None:
						value = str(value).strip()
					if value and (i in idx_lookup):
						self._data[sheet][-1][idx_lookup[i]] = DString(value)
		
		return True
	
	def sheets(self):
		
		return list(self._columns.keys())
	
	def column_count(self, sheet):
		
		if sheet in self._columns:
			return len(self._columns[sheet])
		return []
	
	def row_count(self, sheet):
		
		if sheet not in self._data:
			return 0
		return len(self._data[sheet])
	
	def column_name(self, sheet, column_idx):
		
		if sheet not in self._data:
			return None
		if column_idx not in self._columns[sheet]:
			return None
		return self._columns[sheet][column_idx]
	
	def data(self, sheet, row_idx, column_idx):
		
		if sheet not in self._data:
			return None
		if row_idx > len(self._data[sheet]) - 1:
			return None
		if column_idx not in self._data[sheet][row_idx]:
			return None
		return self._data[sheet][row_idx][column_idx]
	
	def export_data(self, query):
		
		def get_value(item):
			
			value = item.descriptor
			if value is None:
				return ""
			value = value.label.value
			if value is None:
				return ""
			return value
		
		path = as_path(self.url, check_if_exists = False)
		if path is None:
			return
		
		wb = Workbook()
		wb.guess_types = True
		ws = wb.active
		ws.append(query.columns)
		for i in range(len(query.columns)):
			ws.cell(row = 1, column = i + 1).font = Font(bold = True)
		for row in query:
			ws.append([get_value(row[column]) for column in query.columns])
		wb.save(path)

